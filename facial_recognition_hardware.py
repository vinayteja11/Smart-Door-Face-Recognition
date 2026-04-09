import cv2
import face_recognition
import numpy as np
import os
import time
import requests
from gpiozero import OutputDevice
from picamera2 import Picamera2
from datetime import datetime
from openpyxl import Workbook, load_workbook

# ==========================================================
# 🔐 CONFIGURATION
# ==========================================================
TOKEN = "8351024031:AAHPS7z96OIp5KqK95F0o3fVuFUiFXBCLZ4"
CHAT_ID = "1590962234"

IN1 = OutputDevice(14) #
IN2 = OutputDevice(15) #

def stop_motor():
    IN1.off()
    IN2.off()

def unlock_door():
    print("🔓 Unlocking Door...")
    IN1.on()
    IN2.off()
    time.sleep(2)
    stop_motor()

# ==========================================================
# 📊 TRACKING & LOGGING
# ==========================================================
seen_known_ids = set() 
last_unknown_face_encodings = []
unique_male_count = 0
unique_female_count = 0
unique_unknown_count = 0
EXCEL_FILE = "visitor_log.xlsx"

def log_to_excel(name, gender, roll, status):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Time", "Name", "Gender", "Roll No", "Status"])
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    now = datetime.now()
    ws.append([now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"), name, gender, roll, status])
    wb.save(EXCEL_FILE)

# ==========================================================
# 👤 LOAD KNOWN FACES
# ==========================================================
known_encodings = []
known_metadata = [] 

for file in os.listdir("known_faces"):
    if file.endswith((".jpg", ".png")):
        img = face_recognition.load_image_file(f"known_faces/{file}")
        enc = face_recognition.face_encodings(img)
        if enc:
            known_encodings.append(enc[0])
            known_metadata.append(os.path.splitext(file)[0].split("_"))

# ==========================================================
# 📷 CAMERA SETUP
# ==========================================================
picam2 = Picamera2()
config = picam2.create_preview_configuration(main={"size": (640, 480)})
picam2.configure(config)
picam2.start()
time.sleep(2)

# ==========================================================
# 🎯 MAIN LOOP (HIGH FPS VERSION)
# ==========================================================
process_this_frame = True # Optimization: Skip frames
UNKNOWN_COOLDOWN = 15
last_alert_time = 0

while True:
    raw_frame = picam2.capture_array()
    frame = cv2.cvtColor(raw_frame, cv2.COLOR_RGB2BGR) # Color Fix
    
    # Resize to 1/4 size for faster processing
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb_small = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

    if process_this_frame:
        face_locations = face_recognition.face_locations(rgb_small, model="hog")
        face_encodings = face_recognition.face_encodings(rgb_small, face_locations)

        display_list = []
        for face_encoding, face_loc in zip(face_encodings, face_locations):
            name, gender, roll = "Unknown", "N/A", "N/A"
            is_recognized = False
            color = (0, 0, 255)

            if len(known_encodings) > 0:
                distances = face_recognition.face_distance(known_encodings, face_encoding)
                # ADJUSTED THRESHOLD: 0.5 for better recognition
                if np.min(distances) < 0.50:
                    is_recognized = True
                    meta = known_metadata[np.argmin(distances)]
                    name, gender, roll = meta[0], meta[1], meta[2]
                    color = (0, 255, 0)

                    person_id = f"{name}_{roll}"
                    if person_id not in seen_known_ids:
                        seen_known_ids.add(person_id)
                        if gender.lower() == "male": unique_male_count += 1
                        elif gender.lower() == "female": unique_female_count += 1
                        log_to_excel(name, gender, roll, "Authorized")
                        unlock_door()

            if not is_recognized:
                is_new_unknown = True
                if len(last_unknown_face_encodings) > 0:
                    unkn_dist = face_recognition.face_distance(last_unknown_face_encodings, face_encoding)
                    if np.min(unkn_dist) < 0.5: is_new_unknown = False
                
                if is_new_unknown:
                    unique_unknown_count += 1
                    last_unknown_face_encodings.append(face_encoding)
                    log_to_excel("Unknown", "N/A", "N/A", "Denied")

            display_list.append({"loc": face_loc, "info": f"{name} | {gender}", "color": color})

    # Skip next frame to save CPU
    process_this_frame = not process_this_frame

    # --- HUD ---
    cv2.putText(frame, datetime.now().strftime("%H:%M:%S"), (500, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
    cv2.rectangle(frame, (10, 10), (180, 100), (40, 40, 40), -1)
    cv2.putText(frame, f"M: {unique_male_count} F: {unique_female_count}", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
    cv2.putText(frame, f"Unknown: {unique_unknown_count}", (20, 75), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)

    for face in display_list:
        # Scale locations back up by 4x
        top, right, bottom, left = [v * 4 for v in face["loc"]]
        cv2.rectangle(frame, (left, top), (right, bottom), face["color"], 2)
        cv2.putText(frame, face["info"], (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, face["color"], 2)

    cv2.imshow("Smart Door Lock", frame)
    if cv2.waitKey(1) & 0xFF == ord("q"): break

cv2.destroyAllWindows()
picam2.stop()
stop_motor()